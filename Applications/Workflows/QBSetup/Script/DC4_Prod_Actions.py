import openpyxl
from Utilites.Login import Login
from selenium import webdriver
from Utilites.SeleniumOperations import SeleniumOperations
from Applications.Workflows.QBSetup.AppResources import ElementLocators
import time
from selenium.webdriver.common.keys import Keys


class DC4_Prod_Actions:
    # Check_Capability_Status
    def __init__(self, task_type, lo, username,v_Browser):

        self.v_input_wb = openpyxl.load_workbook(ElementLocators.INPUT_FILE_PATH)
        self.v_task_type = task_type
        self.v_input_sheet = self.v_input_wb.get_sheet_by_name("Input")
        self.v_input_sheet_maps = self.v_input_wb.get_sheet_by_name("Maps")
        self.v_input_sheet_Adaptor_Data_Type_ID = self.v_input_wb.get_sheet_by_name("Adaptor Data Type ID")
        self.log = lo
        self.v_username = username
        self.v_Browser=v_Browser
        self.so = SeleniumOperations(self.v_task_type, self.v_Browser, self.log)
        self.login_operations_object = Login(self.v_task_type, self.v_Browser, self.v_input_wb, self.log)
        self.so = SeleniumOperations(self.v_task_type, self.v_Browser, self.log)
        self.login_operations_object.login("Launchpad")
        self.login_operations_object.login("DC4 Prod")
        # time.sleep(5)
        # self.v_Browser.get("https://commerce.spscommerce.com/migrator/")
        # time.sleep(10)
        # self.Migrator("960468")
        # time.sleep(10000)

    # def Migrator(self, profile_uid):
    #     # self.v_Browser.get("https://commerce.spscommerce.com/migrator/")
    #     # time.sleep(10)
    #     self.v_Browser.switch_to.frame(0)
    #     time.sleep(5)
    #     self.v_Browser.find_element_by_xpath(".//*[contains(text(),'Search by column')]//select").click()
    #     time.sleep(2)
    #     self.v_Browser.find_element_by_xpath(".//*[contains(text(),'profile_uid')]").click()
    #     # time.sleep(3)
    #     # driver.find_element_by_xpath(".//b").click()
    #     time.sleep(4)
    #     self.v_Browser.find_element_by_xpath(
    #         "html/body/div[1]/div/div/api-interaction/div/div[4]/label/chosen-select/div/a").click()
    #     time.sleep(4)
    #     self.v_Browser.find_element_by_xpath(
    #         "html/body/div[1]/div/div/api-interaction/div/div[4]/label/chosen-select/div/div/div/input").send_keys(
    #         profile_uid)
    #     # driver.find_element_by_xpath("html/body/div[1]/div/div/api-interaction/div/div[4]/label/chosen-select/div/div/div/input").click()
    #     time.sleep(3)
    #     self.v_Browser.find_element_by_xpath(
    #         "html/body/div[1]/div/div/api-interaction/div/div[4]/label/chosen-select/div/div/div/input").clear()
    #     self.v_Browser.find_element_by_xpath(
    #         "html/body/div[1]/div/div/api-interaction/div/div[4]/label/chosen-select/div/div/div/input").send_keys(
    #         profile_uid)
    #     time.sleep(3)
    #     self.v_Browser.find_element_by_xpath(".//ul/li").click()
    #     # self.v_Browser.switch_to.frame(0)
    #     time.sleep(3)
    #     self.v_Browser.find_element_by_xpath(".//*[contains(text(),'Export')]").click()
    #     self.v_Browser.find_element_by_xpath(".//*[contains(text(),'Migrate')]").click()
    #     # self.v_Browser.find_element_by_xpath(".//*[contains(@class,'medium button confirm')]").click()
    #     print("clicked on migrator")


    def Migrator(self, relationship_uid):
        self.v_Browser.get("https://commerce.spscommerce.com/migrator/")
        time.sleep(6)
        self.v_Browser.switch_to.frame(0)
        self.so.click_element_by_xpath(".//*[contains(text(),'Table')]//select")
        self.so.click_element_by_xpath(".//*[contains(text(),'relationship')]")
        self.so.click_element_by_xpath("html/body/div[1]/div/div/api-interaction/div/div[4]/label/chosen-select/div/a")
        self.so.send_text_by_xpath("html/body/div[1]/div/div/api-interaction/div/div[4]/label/chosen-select/div/div/div/input",relationship_uid)
        time.sleep(6)
        self.so.send_text_by_xpath("html/body/div[1]/div/div/api-interaction/div/div[4]/label/chosen-select/div/div/div/input",relationship_uid)
        time.sleep(6)
        self.so.click_element_by_xpath(".//ul/li")
        self.so.click_element_by_xpath(".//*[contains(text(),'Export')]")
        time.sleep(15)
        self.so.click_element_by_xpath(".//*[contains(text(),'Migrate')]")
        time.sleep(15)
        print("clicked on migrator")


    # def DC4_Prod_Login(self):
    #     self.login_operations_object.login("DC4 Prod")

    def get_maps(self,doc_type,adaptor,retailer_version):
        for i in range(1,self.v_input_sheet_maps.max_row+1):
            doc_type_from_sheet = self.v_input_sheet_maps.cell(row=i, column=1).value
            if str(doc_type_from_sheet) == str(doc_type):
                supplier_version_from_sheet = self.v_input_sheet_maps.cell(row=i, column=2).value
                if adaptor == "Quickbooks" or adaptor == "Fishbowl":
                    supplier_version = "7.2"
                if adaptor == "Dwyer" or adaptor == "Peachtree":
                    supplier_version = "7"
                if str(supplier_version_from_sheet) == str(supplier_version):
                    retailer_version_from_sheet = self.v_input_sheet_maps.cell(row=i, column=3).value
                    if str(retailer_version_from_sheet) == str(retailer_version):
                        maps = self.v_input_sheet_maps.cell(row=i, column=5).value
                        arr_maps = maps.split(",")
                        return arr_maps

    def get_capability_name(self,adaptor, doc):
        Quickbooks = {"810": ["SPS QuickBooks Adaptor | RSX 7.2 | 810 - Legacy", "106968"],
                      "850": ["SPS QuickBooks Adaptor | RSX 7.2 | 850 - Legacy", "106967"],
                      "875": ["SPS QuickBooks Adaptor | RSX 7.2 | 875 - Legacy", "110240"],
                      "856": ["SPS Quickbooks Adapter RSX 7.2 | OzLink | 856", "135124"]}
        Fishbowl = {"810": ["SPS Fishbowl Adaptor | RSX 7.2 | 810 - Legacy", "109097"],
                    "850": ["SPS Fishbowl Adaptor | RSX 7.2 | 850 - Legacy", "109095"],
                    "875": ["SPS FISHBOWL ADAPTOR | RSX 7.2 | 875 - LEGACY", "112556"],
                    "856": ["SPS Fishbowl Adaptor | RSX 7.2 | 856 - Legacy", "109096"]}
        Dwyer = {"810": ["Dwyer Adaptor V7 810 XML", "31788"], "850": ["Dwyer Adaptor V7 850 XML", "31768"],
                 "875": ["Dwyer Adaptor V7 875 XML", "70230"]}
        Peachtree = {"810": ["Peachtree 810 XML", "73633"], "850": ["Peachtree 850 XML", "73632"],
                     "875": ["Peachtree 875 XML", "78248"]}
        if adaptor == "Quickbooks":
            arr = Quickbooks.get(doc)
            return arr
        if adaptor == "Fishbowl":
            arr = Fishbowl.get(doc)
            return arr
        if adaptor == "Dwyer":
            arr = Dwyer.get(doc)
            return arr
        if adaptor == "Peachtree":
            arr = Peachtree.get(doc)
            return arr

    def Search_By_TPID(self,TPID):

        # self.so.click_element_by_xpath(ElementLocators.Browse_Customers)
        self.v_Browser.get(ElementLocators.DC4_Prod_link)
        self.so.click_element_by_xpath(ElementLocators.Customers_by_TPID)
        self.so.send_text_by_xpath(ElementLocators.TPID_input_box,TPID)
        self.so.click_element_by_xpath(ElementLocators.Search_button)
        Company_name=self.so.get_text_by_xpath(ElementLocators.Company_name)
        Profile_name=self.so.get_text_by_xpath(ElementLocators.Profile_name)
        return Company_name+"$"+Profile_name

    def Open_supplier(self,TPID):
        self.so.click_element_by_xpath(ElementLocators.First_company_name)
        self.so.click_element_by_xpath(ElementLocators.Relationships)
        self.so.click_element_by_xpath(ElementLocators.Relationships_Advanced)
        if self.so.check_exists_by_xpath(ElementLocators.Sender_as_Show_all):
            self.so.click_element_by_xpath(ElementLocators.Sender_as_Show_all)
            self.so.click_element_by_xpath(ElementLocators.Show_all_profile)
            self.so.click_element_by_xpath(ElementLocators.Receiver_as_Show_all)
            self.so.click_element_by_xpath(ElementLocators.Show_all_profile2)
            time.sleep(5)
        Retailer_name=self.so.get_text_by_xpath("//*[contains(text(),'"+str(TPID)+"') and contains(@id,'form1:table3:')]//preceding::span[5]")
        Relationship_UID_as_Sender=self.so.get_text_by_xpath("//*[contains(text(),'"+str(TPID)+"') and contains(@id,'form1:table3:')]//preceding::span[6]")
        Relationship_UID_as_Receiver=self.so.get_text_by_xpath("//*[contains(text(),'"+str(TPID)+"') and contains(@id,'form1:table4:')]//preceding::span[8]")
        self.so.click_element_by_xpath(ElementLocators.Relationship_Overview)
        self.so.send_text_by_xpath(ElementLocators.Search_TP_box,Retailer_name)
        self.so.click_element_by_xpath(ElementLocators.TP_Search_button)
        return Retailer_name+"$"+Relationship_UID_as_Sender+"$"+Relationship_UID_as_Receiver

    def Open_profile(self,Profile_name):
        Company_EDI_Summary=self.so.get_text_by_xpath(ElementLocators.Company_EDI_Summary)
        Trading_Partner_EDI_Summary=self.so.get_text_by_xpath(ElementLocators.Trading_Partner_EDI_Summary)
        self.so.click_element_by_xpath("//*[contains(text(),'"+str(Profile_name)+"')]")
        self.so.click_element_by_xpath(ElementLocators.Show)
        return Company_EDI_Summary+"$"+Trading_Partner_EDI_Summary


    def Add_New_Capability(self,Doc_Type,capability_ID):
        self.so.click_element_by_xpath(ElementLocators.createCapability)
        time.sleep(3)
        all_windows = self.v_Browser.window_handles
        num_of_windows = len(all_windows)
        print(num_of_windows)
        requested_window = all_windows[1]
        self.v_Browser.switch_to.window(requested_window)
        print(self.v_Browser.current_url)
        print("Total Window: " + str(num_of_windows))
        self.v_Browser.switch_to.frame(0)

        if str(Doc_Type) == '850' or str(Doc_Type) == '875':
            service_UID = '1007'
        if str(Doc_Type) == '810' or str(Doc_Type) == '855' or str(Doc_Type) == '855' or str(Doc_Type) == '856':
            service_UID = '1006'

        self.so.send_text_by_xpath(ElementLocators.service,service_UID)
        self.so.send_text_by_xpath(ElementLocators.Data_Type,"["+capability_ID+"]")
        self.so.click_element_by_xpath(ElementLocators.submit_create_capability)
        self.v_Browser.switch_to.window(self.v_Browser.window_handles[0])

    def add_existing_capability(self,capability_name,capability_ID,Doc_Type):
        self.so.click_element_by_xpath(ElementLocators.addExistingCapability)

        #switch window
        time.sleep(4)
        all_windows = self.v_Browser.window_handles
        num_of_windows = len(all_windows)
        requested_window = all_windows[1]
        self.v_Browser.switch_to.window(requested_window)

        print(self.v_Browser.current_url)
        print("Total Window: "+str(num_of_windows))

        self.v_Browser.switch_to.frame(0)

        # seq = self.v_Browser.find_elements_by_tag_name('frame')
        # print("seq len")
        # print(len(seq))

        # for index in range(len(seq)):
        #     iframe = self.v_Browser.find_elements_by_tag_name('iframe')[index]
        #     print(iframe)

        # i = 0
        time.sleep(9)
        print("------------------")
        count = self.v_Browser.find_elements_by_xpath("//a[contains(text(),'Show')]")
        print(len(count))
        print("------------------")
        # while True:
        status = ''
        for i in range(len(count)):

            required_capability_name = self.so.get_text_by_xpath("//span[@id='form1:table1:" + str(i) + ":outputText6']")
            print(required_capability_name)
            if str(capability_name) == str(required_capability_name):
                print("capability matched...............")
                checkbox = "//input[@id='form1:table1:" + str(i) + ":tableSelectMany1']"
                capability_uid = "//span[@id='form1:table1:" + str(i) + ":outputText2']"
                self.so.click_element_by_xpath(checkbox)
                self.so.click_element_by_xpath(ElementLocators.cap_choose_btn)
                # self.v_Browser.close()
                self.v_Browser.switch_to.window(self.v_Browser.window_handles[0])
                status = 'capability selected'
                break
            else:
                status = 'Capability not available'
        if status == "Capability not available":
            self.v_Browser.close()
            self.v_Browser.switch_to.window(self.v_Browser.window_handles[0])
            self.Add_New_Capability(Doc_Type,capability_ID)
        print(status)


            # print("in while loop")
            # flag = self.so.check_exists_by_xpath(required_capability_name)
            # val=self.so.get_text_by_xpath(required_capability_name)
            # print(val)
            # if self.so.check_exists_by_xpath(required_capability_name):
            #     print("in True loop")
            #     name_from_UI = self.so.get_text_by_xpath(required_capability_name)
            #     if name_from_UI == capability_name:
            #         print("capability matched...............")
            #         checkbox="//input[@id='form1:table1:"+str(i)+":tableSelectMany1']"
            #         capability_uid="//span[@id='form1:table1:"+str(i)+":outputText2']"
            #         self.so.click_element_by_xpath(checkbox)
            #         # self.so.click_element_by_xpath(ElementLocators.choose_btn)
            #         time.sleep(3)
            #         # self.v_Browser.close()
            #         print("about to close")
            #         time.sleep(1190)
            #         break
            #
            # if self.so.check_exists_by_xpath(required_capability_name)==False:
            #     print("in false loop")
            #     time.sleep(3)
            #     self.v_Browser.close()
            #     # requested_window = all_windows[0]
            #     # self.v_Browser.switch_to.window(requested_window)
            #     # self.so.click_element_by_xpath(ElementLocators.Relationships)
            #     time.sleep(1900)
                # capability not available
                # click on
                # break
            # i += 1

        # if flag==False:
        #     break

    def Toggle_profile_capability(self,capability_name):
        self.so.click_element_by_xpath(ElementLocators.Profiles_Tab)
        time.sleep(3)
        i = 0
        action = ''
        while True:
            Datatype_Name = "//span[@id='form1:table1:0:table2:" + str(i) + ":outputText14']"
            flag = self.so.check_exists_by_xpath(Datatype_Name)
            if flag == True:
                name_from_UI = self.so.get_text_by_xpath(Datatype_Name)
                if name_from_UI == capability_name:
                    status = "//span[@id='form1:table1:0:table2:" + str(i) + ":outputText13']"
                    checkbox = "//input[@id='form1:table1:0:table2:" + str(i) + ":tableSelectMany1']	"
                    Extentions = "//a[@id='form1:table1:0:table2:" + str(i) + ":extensionPopup']"
                    checkbox = "//input[@id='form1:table1:0:table2:" + str(i) + ":tableSelectMany1']	"
                    Status = self.so.get_text_by_xpath(status)
                    self.so.click_element_by_xpath(checkbox)
                    self.so.click_element_by_xpath(ElementLocators.Toggle_profile_capability_status)
                    self.so.click_element_by_xpath(ElementLocators.Profiles_Tab)

                    # if Status == "Active":
                    #     action = "do_nothing"
                    #     break
                    # if Status == "Disabled":
                    #     action = "do_nothing"
                        # check exxtentions
                        # self.so.click_element_by_xpath(checkbox)
                    break
            if flag == False:
                # capability not available
                # click on
                # self.add_existing_capability(capability_name, capability_ID, Doc_Type)

                break
            i += 1

    def Add_Extentions(self, capability_name,arr_maps):
        total_extentions = len(arr_maps)
        time.sleep(8)
        self.so.click_element_by_xpath(ElementLocators.Profiles_Tab)
        time.sleep(3)
        i = 0
        action = ''
        while True:
            Datatype_Name = "//span[@id='form1:table1:0:table2:" + str(i) + ":outputText14']"
            flag = self.so.check_exists_by_xpath(Datatype_Name)
            if flag == True:
                name_from_UI = self.so.get_text_by_xpath(Datatype_Name)
                if name_from_UI == capability_name:
                    status = "//span[@id='form1:table1:0:table2:" + str(i) + ":outputText13']"
                    checkbox = "//input[@id='form1:table1:0:table2:" + str(i) + ":tableSelectMany1']	"
                    Extentions = "//a[@id='form1:table1:0:table2:" + str(i) + ":extensionPopup']"
                    checkbox = "//input[@id='form1:table1:0:table2:" + str(i) + ":tableSelectMany1']	"
                    Status = self.so.get_text_by_xpath(status)

                    self.so.click_element_by_xpath(Extentions)
                    time.sleep(4)
                    self.v_Browser.switch_to.window(self.v_Browser.window_handles[1])
                    self.v_Browser.switch_to.frame(0)

                    for i in range(total_extentions):

                        self.so.click_element_by_xpath(ElementLocators.Add_Extention_btn)

                        self.so.click_element_by_xpath(ElementLocators.map_1080)
                        self.so.click_element_by_xpath(ElementLocators.map_1080_choose)

                    for i in range(total_extentions):
                        self.so.click_element_by_xpath("//a[@id='form1:table1dd"+str(i)+"']//img")

                    for i in range(total_extentions):
                        self.so.send_text_by_xpath("//input[@id='form1:table1:"+str(i)+":table2:0:outputText22']",arr_maps[i])

                    self.so.click_element_by_xpath(ElementLocators.Extention_Save_Changes)
                    self.v_Browser.switch_to.window(self.v_Browser.window_handles[0])
                    # if Status == "Active":
                    #     action = "do_nothing"
                    #     break
                    # if Status == "Disabled":
                    #     action = "do_nothing"
                    # check exxtentions
                    # self.so.click_element_by_xpath(checkbox)
                    break
            if flag == False:
                # capability not available
                # click on
                # self.add_existing_capability(capability_name, capability_ID, Doc_Type)

                break
            i += 1




    def Check_Capability_Status(self,capability_name,capability_ID,Doc_Type,arr_maps):
        time.sleep(4)
        i=0
        action=''
        while True:
            Datatype_Name="//span[@id='form1:table1:0:table2:"+str(i)+":outputText14']"
            flag=self.so.check_exists_by_xpath(Datatype_Name)
            if flag==True:
                name_from_UI=self.so.get_text_by_xpath(Datatype_Name)
                if name_from_UI==capability_name:
                    status="//span[@id='form1:table1:0:table2:"+str(i)+":outputText13']"
                    checkbox="//input[@id='form1:table1:0:table2:"+str(i)+":tableSelectMany1']	"
                    Extentions="//a[@id='form1:table1:0:table2:"+str(i)+":extensionPopup']"
                    checkbox="//input[@id='form1:table1:0:table2:"+str(i)+":tableSelectMany1']	"
                    Status=self.so.get_text_by_xpath(status)
                    print(name_from_UI)
                    print(Status)

                    if Status=="Active":
                        action="do_nothing"
                        break
                    if Status=="Disabled":
                        action = "do_nothing"
                        #check exxtentions
                        # self.so.click_element_by_xpath(checkbox)
                    break
            if flag==False:
                #capability not available
                #click on
                self.add_existing_capability(capability_name,capability_ID,Doc_Type)
                self.Toggle_profile_capability(capability_name)
                self.Add_Extentions(capability_name,arr_maps)
                break
            i += 1




    # def login_to_sailpoint(self):
    #     print("login_to_sailpoint")
    #     so = SeleniumOperations(self.v_task_type, self.v_Browser, self.log)
    #
    #     self.v_Browser.get(ElementLocators.SALESFORCE_URL)
    #     so.click_element_by_xpath(".//*[@id='idp_section_buttons']/button[2]")
    #     # so.click_element_by_xpath(".//*[@id='cancel_idp_hint']")#Log In with a Different Account link
    #     # so.click_element_by_xpath(".//*[contains(text(),'SailPoint')]")
    #     # so.click_element_by_xpath(ElementLocators.SAILPOINT_BTN)
    #     # self.v_Browser.get("https://iam.spscommerce.com/login/login?spEntityID=https%3A%2F%2Fspscommerce.my.salesforce.com&goto=https%3A%2F%2Fiam-sso.spscommerce.com%2Fsso%2FSSORedirect%2FmetaAlias%2Fspscommerce%2Fidp%3FReqID%3D_2CAAAAXCfPa15ME8wMGcwMDAwMDA0Qzk4AAAA3kV67IWmbs12h6m088CMAf07tJcrGOLzEuC6k0Aq2epzV0oaRieDf2U4LBX0Ve0lPuXoZm0y4nxzD4vTrqr7X4mue0_5qLojuNHUxFCj61_mB3PZQj97LDNHpjxkZunTSBRH90cSn5xydi1kFAyU8PVF4s8BKuzm6j8kJzse7O_CephJERN1e6PoPj0VKwaWkFnvGnr-YNZyL64Uy3nUyttHGaMkHw4OYrepEuTK31x1D7qp0MwTmiK2l4_ZqOdVBw%26index%3Dnull%26acsURL%3Dhttps%253A%252F%252Fspscommerce.my.salesforce.com%253Fso%253D00D300000000bzv%26spEntityID%3Dhttps%253A%252F%252Fspscommerce.my.salesforce.com%26binding%3Durn%253Aoasis%253Anames%253Atc%253ASAML%253A2.0%253Abindings%253AHTTP-POST")
    #     # time.sleep(2)
    #     # self.v_Browser.get("https://iam.spscommerce.com/login/login?spEntityID=https%3A%2F%2Fspscommerce.my.salesforce.com&goto=https%3A%2F%2Fiam-sso.spscommerce.com%2Fsso%2FSSORedirect%2FmetaAlias%2Fspscommerce%2Fidp%3FReqID%3D_2CAAAAXCfPa15ME8wMGcwMDAwMDA0Qzk4AAAA3kV67IWmbs12h6m088CMAf07tJcrGOLzEuC6k0Aq2epzV0oaRieDf2U4LBX0Ve0lPuXoZm0y4nxzD4vTrqr7X4mue0_5qLojuNHUxFCj61_mB3PZQj97LDNHpjxkZunTSBRH90cSn5xydi1kFAyU8PVF4s8BKuzm6j8kJzse7O_CephJERN1e6PoPj0VKwaWkFnvGnr-YNZyL64Uy3nUyttHGaMkHw4OYrepEuTK31x1D7qp0MwTmiK2l4_ZqOdVBw%26index%3Dnull%26acsURL%3Dhttps%253A%252F%252Fspscommerce.my.salesforce.com%253Fso%253D00D300000000bzv%26spEntityID%3Dhttps%253A%252F%252Fspscommerce.my.salesforce.com%26binding%3Durn%253Aoasis%253Anames%253Atc%253ASAML%253A2.0%253Abindings%253AHTTP-POST")
    #     # self.v_Browser.get("https://iam.spscommerce.com/login/login?spEntityID=https%3A%2F%2Fspscommerce.my.salesforce.com&goto=https%3A%2F%2Fiam-sso.spscommerce.com%2Fsso%2FSSORedirect%2FmetaAlias%2Fspscommerce%2Fidp%3FReqID%3D_2CAAAAXCfPa15ME8wMGcwMDAwMDA0Qzk4AAAA3kV67IWmbs12h6m088CMAf07tJcrGOLzEuC6k0Aq2epzV0oaRieDf2U4LBX0Ve0lPuXoZm0y4nxzD4vTrqr7X4mue0_5qLojuNHUxFCj61_mB3PZQj97LDNHpjxkZunTSBRH90cSn5xydi1kFAyU8PVF4s8BKuzm6j8kJzse7O_CephJERN1e6PoPj0VKwaWkFnvGnr-YNZyL64Uy3nUyttHGaMkHw4OYrepEuTK31x1D7qp0MwTmiK2l4_ZqOdVBw%26index%3Dnull%26acsURL%3Dhttps%253A%252F%252Fspscommerce.my.salesforce.com%253Fso%253D00D300000000bzv%26spEntityID%3Dhttps%253A%252F%252Fspscommerce.my.salesforce.com%26binding%3Durn%253Aoasis%253Anames%253Atc%253ASAML%253A2.0%253Abindings%253AHTTP-POST")
    #     so.send_text_by_xpath(ElementLocators.SAILPOINT_USERNAME_TEXTBOX, ElementLocators.SAILPOINT_CREDENTIAL_USENAME)
    #     so.send_text_by_xpath(ElementLocators.SAILPOINT_PASSWORD_TEXTBOX, ElementLocators.SAILPOINT_CREDENTIAL_PASSWORD)
    #     so.click_element_by_xpath(ElementLocators.SAILPOINT_LOGIN_BTN)
    #
    # def open_FS_Team_Queue(self):
    #     print("open_FS_Team_Queue")
    #     time.sleep(5)
    #     self.v_Browser.get("https://spscommerce.my.salesforce.com/")
    #     time.sleep(5)
    #     self.v_Browser.get("https://spscommerce.my.salesforce.com/")
    #     self.v_Browser.get(ElementLocators.QUEUE_URL)
    #     time.sleep(7)
    #
    #     if self.so.check_exists_by_xpath(ElementLocators.CROSS_SWITCH_BOX) == True:
    #         self.so.click_element_by_xpath(ElementLocators.CROSS_SWITCH_BOX)
    #
    #     self.v_Browser.get(ElementLocators.QUEUE_URL)
    #     time.sleep(2)
    #     self.so.click_element_by_xpath(ElementLocators.QUEUE_LIST)
    #     self.so.click_element_by_xpath(ElementLocators.FS_TEAM_QUEUE)
    #     if self.so.check_exists_by_xpath(ElementLocators.GO_BTN)==True:
    #       self.so.click_element_by_xpath(ElementLocators.GO_BTN)
    #
    # def get_all_ticket_info(self):
    #     print("get_all_ticket_info")
    #     all_tickets = self.v_Browser.find_elements_by_xpath(ElementLocators.ALL_CASES)
    #     Case_number_and_salesforce_id_array=[]
    #     for ii in all_tickets:
    #         # get salceforce ID
    #         try:
    #             Salesforce_ID = ii.get_attribute("id").split("_")[0]
    #             Case_number = ii.text
    #             Case_number_and_salesforce_id_array.append(Case_number+"_"+Salesforce_ID)
    #             # self.open_case_in_new_tab(Salesforce_ID)
    #         except:
    #             print("error occured in getting Salesforce_ID")
    #     return Case_number_and_salesforce_id_array
    #
    # def return_caseno_and_salesforceno(self,i,Case_number_and_salesforce_id_array):
    #     Case_number = Case_number_and_salesforce_id_array[i].split("_")[0]
    #     salesforce_id = Case_number_and_salesforce_id_array[i].split("_")[1]
    #     return Case_number,salesforce_id
    #
    #
    # def open_cases_and_get_info(self,Case_number_and_salesforce_id_array):
    #
    #     print("open_cases_and_get_info")
    #     # j=0
    #     # print(Case_number_and_salesforce_id_array)
    #     for i in range(len(Case_number_and_salesforce_id_array)):
    #         print("i value")
    #         print(i)
    #         data=self.return_caseno_and_salesforceno(i,Case_number_and_salesforce_id_array)
    #         # print(Case_number_and_salesforce_id_array)
    #         Case_number=data[0]
    #         salesforce_id=data[1]
    #         print("==========================================")
    #         print(Case_number)
    #         print(salesforce_id)
    #         print("==========================================")
    #         # time.sleep(5)
    #         print(Case_number_and_salesforce_id_array)
    #
    #         # salesforce_id_array_task=Case_number_and_salesforce_id_array[i]
    #         # print(i)
    #         # print(salesforce_id)
    #         self.v_Browser.get("https://spscommerce.my.salesforce.com/"+str(salesforce_id))
    #         time.sleep(1)
    #         subject=self.v_Browser.find_element_by_xpath(".//*[contains(@id,'Subject')]").get_attribute("value").lower()
    #
    #         # subject=self.so.get_text_by_xpath(".//*[contains(@id,'Subject')]").lower()
    #         # print(subject)
    #         # if "document processing error" in subject:
    #         #     findings="Document Processing Error"
    #         # elif "document error" in subject:
    #         #     findings = "Document Processing Error for document error"
    #         # elif "processing error" in subject:
    #         #     findings = "Document Processing Error for processing error"
    #         # else:
    #         #     findings="Unknown Error"
    #
    #         mail_count=len(self.v_Browser.find_elements_by_xpath(".//*[contains(@class,'feeditemaux cxfeeditemaux EmailMessageAuxBody')]"))
    #
    #         all_mails = self.v_Browser.find_elements_by_xpath(".//*[contains(@class,'feeditemaux cxfeeditemaux EmailMessageAuxBody')]")
    #         # Case_number_and_salesforce_id_array = []
    #         mail1=" "
    #         mail2=" "
    #         mail3=" "
    #         mail4=" "
    #         mail5=" "
    #         mail6 = " "
    #         mail7 = " "
    #         mail8 = " "
    #         mail9 = " "
    #         mail10 = " "
    #         mail11 = " "
    #         mail12 = " "
    #         mail13 = " "
    #         mail14 = " "
    #         mail15 = " "
    #         mail16 = " "
    #         mail17 = " "
    #         mail18 = " "
    #         mail19 = " "
    #         mail20 = " "
    #         mail21 = " "
    #         mail22 = " "
    #         mail23 = " "
    #         mail24 = " "
    #         mail25 = " "
    #
    #         k=0
    #         for ii in all_mails:
    #             k=k+1
    #             mail_text=ii.text
    #             if k==1:
    #                 mail1=mail_text
    #             elif k==2:
    #                 mail2=mail_text
    #             elif k==3:
    #                 mail3=mail_text
    #             elif k==4:
    #                 mail4=mail_text
    #             elif k==5:
    #                 mail5=mail_text
    #             elif k==6:
    #                 mail6=mail_text
    #             elif k==7:
    #                 mail7=mail_text
    #             elif k==8:
    #                 mail8=mail_text
    #             elif k==9:
    #                 mail9=mail_text
    #             elif k==10:
    #                 mail10=mail_text
    #             elif k==11:
    #                 mail11=mail_text
    #             elif k==12:
    #                 mail12=mail_text
    #             elif k==13:
    #                 mail13=mail_text
    #             elif k==14:
    #                 mail14=mail_text
    #             elif k==15:
    #                 mail15=mail_text
    #             elif k==16:
    #                 mail16=mail_text
    #             elif k==17:
    #                 mail17=mail_text
    #             elif k==18:
    #                 mail18=mail_text
    #             elif k==19:
    #                 mail19=mail_text
    #             elif k==20:
    #                 mail20=mail_text
    #             elif k==21:
    #                 mail21=mail_text
    #             elif k==22:
    #                 mail22=mail_text
    #             elif k==23:
    #                 mail23=mail_text
    #             elif k==24:
    #                 mail24=mail_text
    #             elif k==25:
    #                 mail25=mail_text
    #
    #         self.save_findings_in_excel(i,Case_number,salesforce_id,"findings",subject,mail1,mail2,mail3,mail4,mail5)
    #
    #
    # def save_findings_in_excel(self,i,Case_number,salesforce_id,findings,subject,mail1=None,mail2=None,mail3=None,mail4=None,mail5=None,mail6=None,mail7=None,mail8=None,mail9=None,mail10=None,mail11=None,mail12=None,mail13=None,mail14=None,mail15=None,mail16=None,mail17=None,mail18=None,mail19=None,mail20=None,mail21=None,mail22=None,mail23=None,mail24=None,mail25=None):
    #     row_value=i+2
    #     self.v_input_sheet.cell(row=row_value, column=1).value = Case_number
    #     self.v_input_sheet.cell(row=row_value, column=2).value = salesforce_id
    #     self.v_input_sheet.cell(row=row_value, column=3).value = findings
    #     self.v_input_sheet.cell(row=row_value, column=4).value = subject
    #     self.v_input_sheet.cell(row=row_value, column=5).value = mail1
    #     self.v_input_sheet.cell(row=row_value, column=6).value = mail2
    #     self.v_input_sheet.cell(row=row_value, column=7).value = mail3
    #     self.v_input_sheet.cell(row=row_value, column=8).value = mail4
    #     self.v_input_sheet.cell(row=row_value, column=9).value = mail5
    #     self.v_input_sheet.cell(row=row_value, column=10).value = mail6
    #     self.v_input_sheet.cell(row=row_value, column=9).value = mail7
    #     self.v_input_sheet.cell(row=row_value, column=9).value = mail8
    #     self.v_input_sheet.cell(row=row_value, column=9).value = mail9
    #     self.v_input_sheet.cell(row=row_value, column=9).value = mail10
    #     self.v_input_sheet.cell(row=row_value, column=9).value = mail11
    #     self.v_input_sheet.cell(row=row_value, column=9).value = mail12
    #     self.v_input_sheet.cell(row=row_value, column=9).value = mail13
    #     self.v_input_sheet.cell(row=row_value, column=9).value = mail14
    #     self.v_input_sheet.cell(row=row_value, column=9).value = mail15
    #     self.v_input_sheet.cell(row=row_value, column=9).value = mail16
    #     self.v_input_sheet.cell(row=row_value, column=9).value = mail17
    #     self.v_input_sheet.cell(row=row_value, column=9).value = mail18
    #     self.v_input_sheet.cell(row=row_value, column=9).value = mail19
    #     self.v_input_sheet.cell(row=row_value, column=9).value = mail20
    #     self.v_input_sheet.cell(row=row_value, column=9).value = mail21
    #     self.v_input_sheet.cell(row=row_value, column=9).value = mail22
    #     self.v_input_sheet.cell(row=row_value, column=9).value = mail23
    #     self.v_input_sheet.cell(row=row_value, column=9).value = mail24
    #     self.v_input_sheet.cell(row=row_value, column=9).value = mail25
    #
    #
    #     self.v_input_wb.save(ElementLocators.INPUT_FILE_PATH)
    #
    #
    #
    #
    #

